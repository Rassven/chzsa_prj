import openpyxl  # https://openpyxl.readthedocs.io/en/stable
from webservice.models import *


def write_table(work_model, sheet_model, header_dict, header_row_position):
    # Запись заголовков таблицы.
    column_counter = 0
    for key in header_dict.keys():
        cell_row = str(header_row_position)
        cell_column = "ABCDEFGHIJKLMNOPQRSTUVWXYZ"[column_counter]
        sheet_model[cell_column + cell_row].value = key
        sheet_model.column_dimensions[cell_column].width = len(str(key))
        column_counter += 1
    # Запись данных таблицы.
    records_counter = 1
    for cls_obj in work_model:
        row_counter = records_counter + header_row_position  # Стартовая строка заполнения таблицы.
        column_counter = 0  # "A"!
        for value in header_dict.values():
            cell_row = str(row_counter)
            cell_column = "ABCDEFGHIJKLMNOPQRSTUVWXYZ"[column_counter]
            if hasattr(cls_obj, str(value)):
                cls_obj_value = getattr(cls_obj, value)
                sheet_model[cell_column + cell_row].value = cls_obj_value
                if sheet_model.column_dimensions[cell_column].width < len(str(cls_obj_value)):
                    sheet_model.column_dimensions[cell_column].width = len(str(cls_obj_value))
            else:
                if column_counter == 0:
                    sheet_model[cell_column + cell_row].value = records_counter
            column_counter += 1
        records_counter += 1


def create_excel_report(file_name="project_report.xlsx"):
    """Выгрузка данных моделей Machine, Maintenance и ClaimInfo на листы Excel (СУЩЕСТВУЮЩИЙ БУДЕТ ПЕРЕЗАПИСАН!)"""
    book = openpyxl.Workbook()
    book.remove(book.active)  # remove 'Sheet'

    # Machine sheet create:
    sheet_1 = book.create_sheet("машины", 0)  # Set to active
    machine_header_dict = {
        "№ п/п": "",
        "Модель техники": "machine_model",
        "Зав. № машины": "machine_number",
        "Модель двигателя": "engine_model",
        "Зав. № двигателя": "engine_number",
        "Модель трансмиссии (производитель, артикул)": "transmission_model",
        "Зав. № трансмиссии": "transmission_number",
        "Модель ведущего моста": "drive_bridge_model",
        "Зав. № ведущего моста": "drive_bridge_number",
        "Модель управляемого моста": "controlled_bridge_model",
        "Зав. № управляемого моста": "controlled_bridge_number",
        "Договор поставки №": "info_about_delivery_agreement",
        "Дата отгрузки с завода": "shipment_date",
        "Покупатель": "",
        "Грузополучатель (конечный потребитель)": "end_user",
        "Адрес поставки (эксплуатации)": "end_user_address",
        "Комплектация (доп. опции)": "package_contents",
        "Сервисная компания": "service_company"}
    header_row_position = 3  # Номер строки с заголовками. Внимательно! Строки с 1, а колонки с 0 ("A")!
    write_table(Machine, sheet_1, machine_header_dict, header_row_position)
    # sheet_1.auto_filter.ref = "B3:V3"

    # Maintenance sheet create:
    sheet_2 = book.create_sheet("ТО output")  # "Шапка" (1 строка, с сортировкой) не "закреплена".
    maintenance_header_dict = {
        "№ п/п": "",
        "Зав. № машины": "machine_number",
        "Вид ТО": "maintenance_type",
        "Дата проведения ТО": "maintenance_date",
        "Наработка, мото-час": "worked_hours_time",
        "№ заказ-наряда": "order_number",
        "Дата заказ-наряда": "order_date",
        "Организация, проводившая ТО": "maintenance_organization",
        "*Сервисная компания": "service_company"}
    header_row_position = 1  # Номер строки с заголовками. Внимательно! Строки с 1, а колонки с 0 ("A")!
    write_table(MaintenanceInfo, sheet_2, maintenance_header_dict, header_row_position)
    sheet_2.auto_filter.ref = "B1:I1"

    # ClaimInfo sheet create:
    sheet_3 = book.create_sheet("рекламация output")  # "Шапка" (строки: 1-пустая, 2-просто текст) не "закреплена".
    claiminfo_header_dict = {
        "№ п/п": "",
        "Зав. № машины": "machine_number",
        "Дата отказа": "failure_date",
        "Наработка, мото-час": "worked_hours_time",
        "Узел отказа": "failure_node",
        "Описание отказа": "failure_description",
        "Способ восстановления": "recovery_method",
        "Используемые запасные части": "spare_parts",
        "Дата восстановления": "recovery_date",
        "Время простоя техники": "",
        "*Сервисная компания": "service_company"}
    header_row_position = 2  # Номер строки с заголовками. Внимательно! Строки с 1, а колонки с 0 ("A")!
    write_table(ClaimInfo, sheet_3, claiminfo_header_dict, header_row_position)

    book.save(file_name)
    book.close()
