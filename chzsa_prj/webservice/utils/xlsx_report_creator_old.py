import openpyxl  # https://openpyxl.readthedocs.io/en/stable
from webservice.models import *
# OverWrite File!!!


def create_excel_report(file_name="project_report.xlsx"):
    """Выгрузка данных моделей Machine, Maintenance и ClaimInfo на листы файла Excel"""
    book = openpyxl.Workbook()
    book.remove(book.active)  # remove 'Sheet'
    machine_sheet_name = "машины"
    sheet_1 = book.create_sheet(machine_sheet_name, 0)
    machine_header_dict = {
        "№ п/п": "",
        "Модель техники": "machine_model",
        "Зав. № машины": "machine_number",
        "Модель двигателя": "engine_model",
        "Зав. № двигателя": "engine_number",
        "Модель трансмиссии (производитель, артикул)": "transmission_model",
        "Зав. № трансмиссии": "transmission_number",
        "Модель ведущего моста": "drive_bridge_model",
        "Зав. № ведущего моста": "drive_bridge_number",
        "Модель управляемого моста": "controlled_bridge_model",
        "Зав. № управляемого моста": "controlled_bridge_number",
        "Договор поставки №": "info_about_delivery_agreement",
        "Дата отгрузки с завода": "shipment_date",
        "Покупатель": "buyer",
        "Грузополучатель (конечный потребитель)": "end_user",
        "Адрес поставки (эксплуатации)": "end_user_address",
        "Комплектация (доп. опции)": "package_contents",
        "Сервисная компания": "service_company"}
    # Create header
    column_counter = 0
    row_counter = 3  # Номер строки с заголовками. Внимательно! Строки с 1, а колонки с 0 ("A")!
    for key in machine_header_dict.keys():
        cell_row = str(row_counter)
        cell_column = "ABCDEFGHIJKLMNOPQRSTUVWXYZ"[column_counter]
        sheet_1[cell_column + cell_row].value = key
        sheet_1.column_dimensions[cell_column].width = len(str(key))
        column_counter += 1

    records_counter = 1
    for cls_obj in Machine:
        row_counter = records_counter + 3  # Стартовая строка заполнения таблицы.
        column_counter = 0  # "A"!
        for value in machine_header_dict.values():
            cell_row = str(row_counter)
            cell_column = "ABCDEFGHIJKLMNOPQRSTUVWXYZ"[column_counter]
            if hasattr(cls_obj, str(value)):
                cls_obj_value = getattr(cls_obj, value)
                sheet_1[cell_column + cell_row].value = cls_obj_value
                if sheet_1.column_dimensions[cell_column].width < len(str(cls_obj_value)):
                    sheet_1.column_dimensions[cell_column].width = len(str(cls_obj_value))
            else:
                if column_counter == 0:
                    sheet_1[cell_column + cell_row].value = records_counter
            column_counter += 1
        records_counter += 1
    # sheet_1.auto_filter.ref = "B3:V3"

    # Maintenance sheet create:
    maintenance_sheet_name = "ТО output"
    sheet_2 = book.create_sheet(maintenance_sheet_name)  # "Шапка" (1 строка, с сортировкой) не "закреплена".
    maintenance_header_dict = {
        "№ п/п": "",
        "Зав. № машины": "machine_number",
        "Вид ТО": "maintenance_type",
        "Дата проведения ТО": "maintenance_date",
        "Наработка, мото-час": "worked_hours_time",
        "№ заказ-наряда": "order_number",
        "Дата заказ-наряда": "order_date",
        "Организация, проводившая ТО": "maintenance_organization",
        "*Сервисная компания": "service_company"}
    # Create header
    column_counter = 0
    row_counter = 1  # Номер строки с заголовками.
    for key in maintenance_header_dict.keys():
        cell_row = str(row_counter)
        cell_column = "ABCDEFGHIJKLMNOPQRSTUVWXYZ"[column_counter]
        sheet_2[cell_column + cell_row].value = key
        sheet_2.column_dimensions[cell_column].width = len(str(key))
        column_counter += 1

    records_counter = 1
    for cls_obj in Maintenance:
        row_counter = records_counter + 1  # Стартовая строка данных в таблице.
        column_counter = 0
        for key, value in maintenance_header_dict.items():
            cell_row = str(row_counter)
            cell_column = "ABCDEFGHIJKLMNOPQRSTUVWXYZ"[column_counter]
            if hasattr(cls_obj, str(value)):
                cls_obj_value = getattr(cls_obj, value)
                sheet_2[cell_column + cell_row].value = cls_obj_value
                if sheet_2.column_dimensions[cell_column].width < len(str(cls_obj_value)):
                    sheet_2.column_dimensions[cell_column].width = len(str(cls_obj_value))
            else:
                if column_counter == 0:
                    sheet_2[cell_column + cell_row].value = records_counter
            column_counter += 1
        records_counter += 1

    sheet_2.auto_filter.ref = "B1:I1"

    # ClaimInfo sheet create:
    claiminfo_sheet_name = "рекламация output"
    sheet_3 = book.create_sheet(claiminfo_sheet_name)  # "Шапка" (строки: 1-пустая, 2-просто текст) не "закреплена".
    claiminfo_header_dict = {
        "№ п/п": "",
        "Зав. № машины": "machine_number",
        "Дата отказа": "failure_date",
        "Наработка, мото-час": "worked_hours_time",
        "Узел отказа": "failure_node",
        "Описание отказа": "failure_description",
        "Способ восстановления": "recovery_method",
        "Используемые запасные части": "spare_parts",
        "Дата восстановления": "recovery_date",
        "Время простоя техники": "",
        "*Сервисная компания": "service_company"}

    # Create header
    column_counter = 0
    row_counter = 2  # Номер строки с заголовками.
    for key in claiminfo_header_dict.keys():
        cell_row = str(row_counter)
        cell_column = "ABCDEFGHIJKLMNOPQRSTUVWXYZ"[column_counter]
        sheet_3[cell_column + cell_row].value = key
        sheet_3.column_dimensions[cell_column].width = len(str(key))
        column_counter += 1

    records_counter = 1
    for cls_obj in ClaimInfo:
        row_counter = records_counter + 2  # Стартовая строка данных в таблице.
        column_counter = 0
        for value in claiminfo_header_dict.values():
            cell_row = str(row_counter)
            cell_column = "ABCDEFGHIJKLMNOPQRSTUVWXYZ"[column_counter]
            if hasattr(cls_obj, str(value)):
                cls_obj_value = getattr(cls_obj, value)
                sheet_3[cell_column + cell_row].value = cls_obj_value
                if sheet_3.column_dimensions[cell_column].width < len(str(cls_obj_value)):
                    sheet_3.column_dimensions[cell_column].width = len(str(cls_obj_value))
            else:
                if column_counter == 0:
                    sheet_3[cell_column + cell_row].value = records_counter
            column_counter += 1
        records_counter += 1

        book.save(file_name)
        book.close()
